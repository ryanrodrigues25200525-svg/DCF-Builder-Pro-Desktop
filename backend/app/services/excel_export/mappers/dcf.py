from __future__ import annotations
import re
from copy import copy
from datetime import date, datetime
from typing import Any
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from .utils import (
    _sheet, _to_float, _safe_set, _safe_set_or_clear, _display_company_label,
    _sanitize_wacc_rate, _sanitize_terminal_growth_rate, _force_set, _scale, _safe_date,
    _fiscal_year_end_date, _historical_value_for_year, _axis_from_bounds, _numeric_list,
    _matrix_values, _fallback_wacc_terminal_matrix, _fallback_revenue_ebit_matrix,
    _set_percent_axis_row, _set_percent_axis_column, _clear_sensitivity_blocks,
    _scenario_snapshot, _apply_scenario_snapshot_to_sheet, _metric_series,
    _split_cost_of_revenue_components, _opex_component_series, _safe_year_end_date,
    _scenario_assumption_value, _infer_revenue_growth_rate, _link_dcf_income_statement_to_recalculated_data,
    _normalize_public_dcf_assumption_block, _enforce_core_public_dcf_formulas,
    _enforce_outputs_bridge_formulas, _set_comment,
    SHEET_WACC, TEN_YEAR_COLUMNS, RECALC_COLUMNS, DCF_TIMELINE_COLUMNS,
    MAX_TERMINAL_GROWTH_RATE, MIN_TERMINAL_WACC_SPREAD,
    DCF_HELPER_CASH_CELL, DCF_HELPER_DEBT_CELL, DCF_HELPER_NON_OP_CELL,
    DCF_HELPER_CASH_CELL_ABS, DCF_HELPER_DEBT_CELL_ABS, DCF_HELPER_NON_OP_CELL_ABS,
    SHEET_OUTPUTS, SHEET_ASSUMPTION_BREAKDOWN, SHEET_DATA_ORIGINAL
)

SCENARIO_BASE = "base"
SCENARIO_BULL = "bull"
SCENARIO_BEAR = "bear"

def _map_dcf_base_inputs(dcf_base: Worksheet, payload: dict[str, Any], divisor: float) -> None:
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    market = payload.get("market", {})
    market = market if isinstance(market, dict) else {}
    transaction = payload.get("transaction", {})
    transaction = transaction if isinstance(transaction, dict) else {}
    company = payload.get("company", {})
    company = company if isinstance(company, dict) else {}
    forecasts = payload.get("forecasts", []) or []
    key_metrics = (payload.get("uiMeta") or {}).get("keyMetrics") or {}

    tax_rate = _to_float(assumptions.get("taxRate"))
    da_pct = _to_float(assumptions.get("daPctRevenue"))
    if da_pct is None:
        da_pct = _to_float(assumptions.get("deaRatio"))
    terminal_assumptions = assumptions.get("terminal") or {}
    exit_multiple = _to_float(terminal_assumptions.get("exitMultiple"))
    wacc_assumptions = assumptions.get("wacc")
    wacc_assumptions = wacc_assumptions if isinstance(wacc_assumptions, dict) else {}
    base_wacc_assumption = _sanitize_wacc_rate(assumptions.get("waccRate") or wacc_assumptions.get("waccRate"))
    terminal_growth = _sanitize_terminal_growth_rate(
        terminal_assumptions.get("g"),
        reference_wacc=base_wacc_assumption,
    )
    revenue_growth = _to_float(assumptions.get("revenueGrowth"))
    if revenue_growth is None:
        revenue_growth = _to_float(assumptions.get("revenueGrowthRate"))
    if revenue_growth is None:
        revenue_growth = _infer_revenue_growth_rate(forecasts)
    if revenue_growth is None:
        revenue_growth = 0.06

    shares = _to_float(market.get("sharesDiluted")) or 0.0
    price = _to_float(market.get("currentPrice")) or 0.0
    market_cap_raw = _to_float(market.get("marketCap"))
    if market_cap_raw is None and shares > 0 and price > 0:
        market_cap_raw = shares * price

    debt_raw = _to_float(market.get("marketValueDebt"))
    if debt_raw is None:
        debt_raw = _to_float(market.get("debt")) or 0.0
    cash_raw = _to_float(market.get("cash")) or 0.0
    non_operating_assets_raw = _to_float(market.get("nonOperatingAssets")) or 0.0

    equity_raw = _to_float(key_metrics.get("equityValue"))
    if equity_raw is None:
        equity_raw = market_cap_raw

    net_debt_raw = _to_float(market.get("netDebt"))
    if net_debt_raw is None:
        net_debt_raw = debt_raw - cash_raw

    explicit_purchase_price = None
    if isinstance(transaction, dict):
        explicit_purchase_price = _to_float(transaction.get("purchasePrice"))
    if explicit_purchase_price is None:
        explicit_purchase_price = _to_float(payload.get("purchasePrice"))

    enterprise_raw = explicit_purchase_price if explicit_purchase_price is not None else _to_float(key_metrics.get("enterpriseValue"))
    if enterprise_raw is None and equity_raw is not None and net_debt_raw is not None:
        enterprise_raw = equity_raw + net_debt_raw

    _safe_set(dcf_base, "C9", _scale(enterprise_raw, divisor))
    _safe_set(dcf_base, "C11", _scale(equity_raw, divisor))
    _safe_set(dcf_base, "F16", _scale(net_debt_raw, divisor))
    _safe_set(dcf_base, "F17", _scale(cash_raw, divisor))
    _safe_set(dcf_base, "F18", _scale(debt_raw, divisor))
    _safe_set(dcf_base, "F19", _scale(non_operating_assets_raw, divisor))

    capex = None
    nwc_change = None

    first_forecast = forecasts[0] if forecasts and isinstance(forecasts[0], dict) else {}
    capex = _to_float(first_forecast.get("capex"))
    nwc_change = _to_float(first_forecast.get("nwcChange"))

    if capex is None:
        capex_abs = assumptions.get("capexAbsolute")
        if isinstance(capex_abs, list) and capex_abs:
            capex = _to_float(capex_abs[0])

    if capex is not None:
        capex = abs(capex)
    if nwc_change is not None:
        nwc_change = abs(nwc_change)

    _safe_set(dcf_base, "F9", _scale(capex, divisor))
    _safe_set(dcf_base, "F10", _scale(nwc_change, divisor))
    _safe_set(dcf_base, "F11", tax_rate)
    _safe_set(dcf_base, "F13", da_pct)
    _safe_set(dcf_base, "F14", revenue_growth)
    dcf_base["F14"].number_format = "0.0%"
    _safe_set(dcf_base, "C16", exit_multiple)
    _safe_set(dcf_base, "Q103", terminal_growth)

    _set_comment(dcf_base, "F9", "Source: Forecast capex (payload.forecasts[0].capex) or capexAbsolute fallback.")
    _set_comment(dcf_base, "F10", "Source: Forecast working-capital change (payload.forecasts[0].nwcChange).")
    _set_comment(dcf_base, "F11", "Source: Tax assumption from payload.assumptions.taxRate.")
    _set_comment(dcf_base, "F14", "Source: Revenue growth assumption; defaults to inferred forecast growth.")
    _set_comment(dcf_base, "C16", "Source: Terminal exit multiple from payload.assumptions.terminal.exitMultiple.")
    _set_comment(dcf_base, "F17", "Source: Market cash and equivalents.")
    _set_comment(dcf_base, "F18", "Source: Market debt / market value debt proxy.")
    _set_comment(dcf_base, "F19", "Source: Non-operating assets from market payload.")

    as_of = _safe_date(company.get("asOfDate"))
    if as_of is not None:
        _safe_set(dcf_base, "I9", as_of)

    fiscal_year = as_of.year if as_of is not None else datetime.now().year
    fiscal_end = _fiscal_year_end_date(company.get("fiscalYearEnd"), fiscal_year)
    if fiscal_end is None:
        fiscal_end = date(fiscal_year, 12, 31)
    _safe_set(dcf_base, "I11", fiscal_end)



def _sync_shared_scenario_inputs(scenario_sheet: Worksheet, dcf_base: Worksheet, *, nwc_multiplier: float) -> None:
    # Preserve scenario-specific formulas and apply only payload-driven shared assumptions.
    for cell in ("C9", "C11", "F9", "F11", "F13", "F14", "C16", "I9", "I11", "Q103"):
        scenario_sheet[cell].value = dcf_base[cell].value

    base_nwc_change = _to_float(dcf_base["F10"].value)
    if base_nwc_change is not None:
        scenario_sheet["F10"].value = base_nwc_change * max(0.0, nwc_multiplier)



def _harden_growth_rate_formulas(*scenario_sheets: Worksheet) -> None:
    # Guard CAGR calculations against divide-by-zero in low-data scenarios.
    formula_map = {
        "S24": "=IFERROR((I24/H24)^(1/(COLUMNS(H24:I24)-1))-1,0)",
        "T24": "=IFERROR((Q24/J24)^(1/(COLUMNS(J24:Q24)-1))-1,0)",
        "S27": "=IFERROR((I27/H27)^(1/(COLUMNS(H27:I27)-1))-1,0)",
        "T27": "=IFERROR((Q27/J27)^(1/(COLUMNS(J27:Q27)-1))-1,0)",
    }
    for sheet in scenario_sheets:
        for cell_ref, formula in formula_map.items():
            _force_set(sheet, cell_ref, formula)
        # Harden any remaining CAGR-style formulas in S/T columns that still
        # divide by historical anchors without IFERROR wrappers.
        for row in range(20, 131):
            for col in ("S", "T"):
                cell_ref = f"{col}{row}"
                value = sheet[cell_ref].value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                upper_value = value.upper()
                if "IFERROR(" in upper_value:
                    continue
                if "COLUMNS(" not in upper_value:
                    continue
                expression = value[1:].lstrip("+")
                _force_set(sheet, cell_ref, f"=IFERROR({expression},0)")



def _normalize_dcf_waterfall_formulas(*scenario_sheets: Worksheet) -> None:
    # Keep sign conventions intuitive while preserving economics:
    # OpEx rows positive, EBIT subtracts OpEx; CapEx/NWC assumptions reference fixed inputs directly.
    for sheet in scenario_sheets:
        for col in DCF_TIMELINE_COLUMNS:
            _force_set(sheet, f"{col}48", f"={col}36+{col}39+{col}42+{col}45")
            _force_set(sheet, f"{col}49", f"=IFERROR({col}48/{col}20,0)")
            _force_set(sheet, f"{col}51", f"={col}32-{col}48")
            _force_set(sheet, f"{col}65", "=-$F$9")
            _force_set(sheet, f"{col}77", "=-$F$10")
            _force_set(sheet, f"{col}69", "=$F$13")



def _normalize_public_dcf_layout(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    _link_dcf_income_statement_to_recalculated_data(dcf_base, dcf_bull, dcf_bear)
    _normalize_public_dcf_assumption_block(dcf_base, dcf_bull, dcf_bear, payload, divisor)
    _enforce_core_public_dcf_formulas(dcf_base, dcf_bull, dcf_bear)
    _enforce_outputs_bridge_formulas(outputs)



def _sync_scenario_formula_backbone(dcf_base: Worksheet, *scenario_sheets: Worksheet) -> None:
    # Keep formula topology identical across Base/Bull/Bear to avoid scenario drift.
    scenario_specific_formula_cells = {"F12"}
    for row in dcf_base.iter_rows(min_row=1, max_row=dcf_base.max_row, min_col=1, max_col=dcf_base.max_column):
        for base_cell in row:
            formula = base_cell.value
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue
            if base_cell.coordinate in scenario_specific_formula_cells:
                continue
            for scenario in scenario_sheets:
                scenario_cell = scenario[base_cell.coordinate]
                if isinstance(scenario_cell, MergedCell):
                    continue
                scenario_cell.value = formula



def _apply_capex_schedule_to_dcf(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    # Projection-period CapEx rows are formula-driven from assumptions.
    # Avoid stamping hardcoded year-by-year values that can drift.
    return



def _apply_scenario_snapshots_to_dcf(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    base_snapshot = _scenario_snapshot(payload, SCENARIO_BASE)
    bull_snapshot = _scenario_snapshot(payload, SCENARIO_BULL)
    bear_snapshot = _scenario_snapshot(payload, SCENARIO_BEAR)

    if base_snapshot is None and bull_snapshot is None and bear_snapshot is None:
        return

    if base_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_base, base_snapshot, timeline_years, divisor)
    if bull_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_bull, bull_snapshot, timeline_years, divisor)
    if bear_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_bear, bear_snapshot, timeline_years, divisor)



def _finalize_assumption_block_cleanup(*scenario_sheets: Worksheet) -> None:
    # Ensure assumption area is clean and free of duplicate helper numbers.
    for sheet in scenario_sheets:
        _safe_set(sheet, "B18", "Income Statement")
        _safe_set_or_clear(sheet, "B19", None)
        _safe_set_or_clear(sheet, "E15", None)
        _safe_set_or_clear(sheet, "E16", None)
        _safe_set_or_clear(sheet, "F16", None)
        _safe_set_or_clear(sheet, "F17", None)
        _safe_set_or_clear(sheet, "F18", None)
        _safe_set_or_clear(sheet, "F19", None)
        _safe_set_or_clear(sheet, "E19", None)
        for address in ("C18", "C19", "D18", "D19", "E18", "G18", "G19"):
            _safe_set_or_clear(sheet, address, None)
        # Remove lingering note indicators (red triangles) from template/input mapping.
        for address in (
            "C13",
            "C16",
            "C17",
            "E15",
            "E16",
            "F9",
            "F10",
            "F11",
            "F12",
            "F13",
            "F14",
            "F15",
            "F16",
            "F17",
            "F18",
            "F19",
        ):
            sheet[address].comment = None



def _add_prior_actual_year_display_column(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    historical_years: set[int],
    divisor: float,
) -> None:
    if not timeline_years:
        return
    prior_candidates = [year for year in historical_years if year < timeline_years[0]]
    if not prior_candidates:
        return

    prior_year = max(prior_candidates)
    prior_label = f"FY{prior_year}A"

    revenue = _historical_value_for_year(payload, prior_year, statement="income", keys=["Total Revenue", "Revenue", "Sales"])
    cost_of_revenue = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["Cost of Revenue", "COGS", "Cost Of Revenue", "Cost of Sales", "Purchases"],
    )
    rnd = _historical_value_for_year(payload, prior_year, statement="income", keys=["Research & Development", "R&D", "Research and Development"])
    sga = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["SG&A", "SGA", "General and Administrative", "GeneralAndAdministrative", "G&A", "GA"],
    )
    da = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["D&A (included in Operating)", "D&A", "DA", "Depreciation & Amortization", "Depreciation"],
    )
    if da is None:
        da = _historical_value_for_year(payload, prior_year, statement="cashflow", keys=["Depreciation"])
    other_opex = _historical_value_for_year(payload, prior_year, statement="income", keys=["Other Operating Expenses", "Other"])
    ebit = _historical_value_for_year(payload, prior_year, statement="income", keys=["Operating Income (EBIT)", "EBIT", "Operating Income"])
    capex = _historical_value_for_year(payload, prior_year, statement="cashflow", keys=["Capex", "Capital Expenditures", "Capital Expenditure"])

    for sheet in (dcf_base, dcf_bull, dcf_bear):
        # Carry timeline header style to the added prior-year display column.
        for row in (18, 63, 72):
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)
            _force_set(sheet, f"G{row}", prior_label)

        # Apply consistent number/border styles from first timeline column.
        for row in (20, 24, 27, 30, 32, 33, 36, 39, 42, 45, 48, 49, 51, 52, 54, 55, 57, 60, 65, 66, 67, 68, 69, 74, 75, 76, 77, 78, 79):
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)

        if revenue is not None:
            _force_set(sheet, "G20", _scale(revenue, divisor))
            _force_set(sheet, "H21", "=IFERROR(H20/G20-1,0)")
        _force_set(sheet, "G21", "-")
        if cost_of_revenue is not None:
            _force_set(sheet, "G24", _scale(cost_of_revenue, divisor))
        _force_set(sheet, "G25", "=IFERROR(G24/G20,0)")
        _force_set(sheet, "G28", "=IFERROR(G27/G20,0)")
        _force_set(sheet, "G30", "=G24+G27")
        _force_set(sheet, "G32", "=G20-G30")
        _force_set(sheet, "G33", "=IFERROR(G32/G20,0)")
        if rnd is not None:
            _force_set(sheet, "G36", _scale(rnd, divisor))
        _force_set(sheet, "G37", "=IFERROR(G36/G20,0)")
        if sga is not None:
            _force_set(sheet, "G39", _scale(sga, divisor))
        _force_set(sheet, "G40", "=IFERROR(G39/G20,0)")
        if da is not None:
            _force_set(sheet, "G42", _scale(da, divisor))
        _force_set(sheet, "G43", "=IFERROR(G42/G20,0)")
        if other_opex is not None:
            _force_set(sheet, "G45", _scale(other_opex, divisor))
        _force_set(sheet, "G46", "=IFERROR(G45/G20,0)")
        _force_set(sheet, "G48", "=G36+G39+G42+G45")
        _force_set(sheet, "G49", "=IFERROR(G48/G20,0)")
        if ebit is not None:
            _force_set(sheet, "G51", _scale(ebit, divisor))
        else:
            _force_set(sheet, "G51", "=G32-G48")
        _force_set(sheet, "G52", "=IFERROR(G51/G20,0)")
        _force_set(sheet, "G54", "=G51+G68")
        _force_set(sheet, "G55", "=IFERROR(G54/G20,0)")
        _force_set(sheet, "G57", "=-G51*$F$11")
        _force_set(sheet, "G58", "=$F$11")
        _force_set(sheet, "G60", "=G51+G57")
        _force_set(sheet, "G61", "=IFERROR(G60/G20,0)")

        if capex is not None:
            _force_set(sheet, "G65", -(_scale(capex, divisor) or 0.0))
        else:
            _force_set(sheet, "G65", "=H65")
        _force_set(sheet, "G66", "=IFERROR(-G65/G20,0)")
        _safe_set_or_clear(sheet, "G67", None)
        if da is not None:
            _force_set(sheet, "G68", _scale(da, divisor))
        else:
            _force_set(sheet, "G68", "=G69*G20")
        _force_set(sheet, "G69", "=$F$13")
        _force_set(sheet, "G70", "=IFERROR(-G68/G65,0)")

        _force_set(sheet, "G74", "=G60")
        _force_set(sheet, "G75", "=G68")
        _force_set(sheet, "G76", "=G65")
        _force_set(sheet, "G77", "=0")
        _force_set(sheet, "G78", "=SUM(G74:G77)")
        _force_set(sheet, "G79", "=IFERROR(G78/G20,0)")


def _map_sensitivity_blocks(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    terminal = assumptions.get("terminal", {})
    terminal = terminal if isinstance(terminal, dict) else {}
    sensitivities = payload.get("sensitivities", {})
    sensitivities = sensitivities if isinstance(sensitivities, dict) else {}

    key_metrics = (payload.get("uiMeta") or {}).get("keyMetrics") or {}
    market = payload.get("market", {})
    shares = _to_float(market.get("sharesDiluted")) or 0.0
    price = _to_float(market.get("currentPrice")) or 0.0
    market_cap = shares * price if shares > 0 and price > 0 else None
    net_debt = _to_float(market.get("netDebt"))
    if net_debt is None:
        debt = _to_float(market.get("debt")) or 0.0
        cash = _to_float(market.get("cash")) or 0.0
        net_debt = debt - cash

    base_ev_raw = _to_float(key_metrics.get("enterpriseValue"))
    if base_ev_raw is None and market_cap is not None:
        base_ev_raw = market_cap + (net_debt or 0.0)
    if base_ev_raw is None:
        base_ev_raw = 1_000_000.0

    pv_terminal_raw = _to_float(key_metrics.get("pvTerminalValue")) or 0.0
    tv_weight = pv_terminal_raw / base_ev_raw if base_ev_raw > 0 else 0.7
    tv_weight = max(0.4, min(0.9, tv_weight))

    base_wacc = _sanitize_wacc_rate(assumptions.get("waccRate")) or 0.10
    base_growth = _sanitize_terminal_growth_rate(terminal.get("g"), reference_wacc=base_wacc) or 0.025
    base_revenue_growth = _to_float(assumptions.get("revenueGrowth")) or 0.03
    base_ebit_margin = _to_float(assumptions.get("ebitMargin")) or _to_float(assumptions.get("ebitMarginTarget")) or 0.15

    wacc_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("waccAxis")) or _numeric_list(sensitivities.get("waccGrid")),
        center=base_wacc,
        step=0.01,
        min_value=0.01,
        max_value=0.30,
    )
    growth_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("terminalGrowthAxis")) or _numeric_list(sensitivities.get("gGrid")),
        center=base_growth,
        step=0.005,
        min_value=0.0,
        max_value=max(0.001, min(0.08, min(wacc_axis) - 0.001)),
    )
    revenue_growth_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("revenueGrowthAxis")),
        center=base_revenue_growth,
        step=0.01,
        min_value=-0.10,
        max_value=0.30,
    )
    ebit_margin_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("ebitMarginAxis")),
        center=base_ebit_margin,
        step=0.01,
        min_value=0.01,
        max_value=0.60,
    )

    wacc_terminal_matrix = _matrix_values(sensitivities.get("waccTerminalEvMatrix"))
    if wacc_terminal_matrix is None:
        wacc_terminal_matrix = _fallback_wacc_terminal_matrix(
            base_ev=base_ev_raw,
            base_wacc=base_wacc,
            base_growth=base_growth,
            wacc_axis=wacc_axis,
            growth_axis=growth_axis,
            tv_weight=tv_weight,
        )

    revenue_ebit_matrix = _matrix_values(sensitivities.get("revenueEbitEvMatrix"))
    if revenue_ebit_matrix is None:
        revenue_ebit_matrix = _fallback_revenue_ebit_matrix(
            base_ev=base_ev_raw,
            base_revenue_growth=base_revenue_growth,
            base_ebit_margin=base_ebit_margin,
            revenue_growth_axis=revenue_growth_axis,
            ebit_margin_axis=ebit_margin_axis,
        )

    for sheet in (dcf_base, dcf_bull, dcf_bear):
        # Add visual space between the two sensitivity tables.
        current_i_width = sheet.column_dimensions["I"].width
        if current_i_width is None or current_i_width < 14:
            sheet.column_dimensions["I"].width = 14

        _safe_set(sheet, "C117", "Enterprise Value - WACC x Terminal Growth")
        _safe_set(sheet, "D118", "WACC")
        _force_set(sheet, "C119", f"='{SHEET_OUTPUTS}'!$D$35")
        _safe_set(sheet, "B121", "Terminal Growth")
        _safe_set(sheet, "B122", "Rate")
        _set_percent_axis_row(sheet, cells=("D119", "E119", "F119", "G119", "H119"), values=wacc_axis)
        _set_percent_axis_column(sheet, cells=("C120", "C121", "C122", "C123", "C124"), values=growth_axis)

        for row_idx in range(5):
            for col_idx in range(5):
                scaled = _scale(wacc_terminal_matrix[row_idx][col_idx], divisor)
                _force_set(sheet, f"{chr(ord('D') + col_idx)}{120 + row_idx}", scaled)

        _force_set(sheet, "C126", "=MIN(D120:H124)")
        _force_set(sheet, "C127", "=PERCENTILE(D120:H124,0.25)")
        _force_set(sheet, "C128", "=MEDIAN(D120:H124)")
        _force_set(sheet, "C129", "=PERCENTILE(D120:H124,0.75)")
        _force_set(sheet, "C130", "=MAX(D120:H124)")
        _safe_set(sheet, "B126", "Min")
        _safe_set(sheet, "B127", "Q1")
        _safe_set(sheet, "B128", "Median")
        _safe_set(sheet, "B129", "Q3")
        _safe_set(sheet, "B130", "Max")

        # Ensure every right-side table cell is materialized in sheet XML.
        for row in range(117, 131):
            for col in range(ord("I"), ord("O") + 1):
                ref = f"{chr(col)}{row}"
                if sheet[ref].value is None:
                    _force_set(sheet, ref, "")

        _safe_set(sheet, "I117", "Enterprise Value - Revenue Growth x EBIT Margin")
        _safe_set(sheet, "J118", "Revenue Growth")
        _safe_set(sheet, "I121", "EBIT Margin")
        _safe_set(sheet, "I122", "Rate")
        _set_percent_axis_row(sheet, cells=("J119", "K119", "L119", "M119", "N119"), values=revenue_growth_axis)
        _set_percent_axis_column(sheet, cells=("I120", "I121", "I122", "I123", "I124"), values=ebit_margin_axis)

        for row_idx in range(5):
            for col_idx in range(5):
                scaled = _scale(revenue_ebit_matrix[row_idx][col_idx], divisor)
                numeric = scaled if scaled is not None else 0.0
                _force_set(sheet, f"{chr(ord('J') + col_idx)}{120 + row_idx}", numeric)

        _safe_set(sheet, "I126", "Min")
        _safe_set(sheet, "I127", "Q1")
        _safe_set(sheet, "I128", "Median")
        _safe_set(sheet, "I129", "Q3")
        _safe_set(sheet, "I130", "Max")
        _force_set(sheet, "J126", "=MIN(J120:N124)")
        _force_set(sheet, "J127", "=PERCENTILE(J120:N124,0.25)")
        _force_set(sheet, "J128", "=MEDIAN(J120:N124)")
        _force_set(sheet, "J129", "=PERCENTILE(J120:N124,0.75)")
        _force_set(sheet, "J130", "=MAX(J120:N124)")

        # Highlight low/median/high valuation outcomes with a standard red-yellow-green gradient.
        for target_range in ("D120:H124", "J120:N124"):
            sheet.conditional_formatting.add(
                target_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    _clear_sensitivity_blocks(dcf_bull, dcf_bear)



def _replace_template_placeholders(
    *,
    company_name: str | None,
    ticker: str,
    sheets: tuple[Worksheet, ...],
) -> None:
    long_name = company_name or ticker
    replacements = {
        "ABC/SNS Health": long_name,
        "ABC Corp.": long_name,
        "ABC Corp": long_name,
        "ABC": ticker,
    }

    for sheet in sheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.data_type == "f":
                    continue
                if not isinstance(cell.value, str):
                    continue

                updated = cell.value
                for old, new in replacements.items():
                    updated = updated.replace(old, new)
                if updated != cell.value:
                    cell.value = updated

def _finalize_timeline_headers(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    timeline_years: list[int],
    historical_years: set[int],
) -> None:
    # Final guardrail: force all timeline headers to explicit FY labels as text.
    for idx, col in enumerate(DCF_TIMELINE_COLUMNS):
        year = timeline_years[idx]
        label = f"FY{year}{'A' if year in historical_years else 'E'}"
        _force_set(outputs, f"{col}6", label)
        outputs[f"{col}6"].number_format = "@"
        for sheet in (dcf_base, dcf_bull, dcf_bear):
            _force_set(sheet, f"{col}18", label)
            _force_set(sheet, f"{col}63", label)
            _force_set(sheet, f"{col}72", label)
            _force_set(sheet, f"{col}89", label)
            sheet[f"{col}89"].number_format = "@"



def _reset_dcf_sheet_view_to_top(*sheets: Worksheet) -> None:
    # Ensure each scenario tab opens at the top of the sheet instead of
    # preserving a template viewport near the bottom.
    for sheet in sheets:
        sheet.sheet_view.topLeftCell = "A1"
        if sheet.sheet_view.selection:
            sheet.sheet_view.selection[0].activeCell = "A1"
            sheet.sheet_view.selection[0].sqref = "A1"



def _remove_assumption_breakdown(workbook: Workbook, cover: Worksheet) -> None:
    if SHEET_ASSUMPTION_BREAKDOWN in workbook.sheetnames:
        workbook.remove(workbook[SHEET_ASSUMPTION_BREAKDOWN])
    if SHEET_DATA_ORIGINAL in workbook.sheetnames:
        workbook.remove(workbook[SHEET_DATA_ORIGINAL])
    if "Data ->" in workbook.sheetnames:
        workbook.remove(workbook["Data ->"])

    # Keep cover TOC coherent after removing helper tabs.
    _safe_set(cover, "F15", "Data Given (Recalculated)")
    _safe_set_or_clear(cover, "F16", None)
    _safe_set_or_clear(cover, "F17", None)



