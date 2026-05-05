from __future__ import annotations
from copy import copy
from typing import Any
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .constants import (
    DCF_TIMELINE_COLUMNS, RECALC_COLUMNS, SHEET_DATA_RECALCULATED,
    SHEET_DCF_BASE, SHEET_DCF_BULL, SHEET_DCF_BEAR,
    DCF_HELPER_CASH_CELL, DCF_HELPER_DEBT_CELL, DCF_HELPER_NON_OP_CELL,
    DCF_HELPER_CASH_CELL_ABS, DCF_HELPER_DEBT_CELL_ABS, DCF_HELPER_NON_OP_CELL_ABS
)
from .core import (
    _to_float, _first_float, _scale
)
from .financial import (
    _sanitize_wacc_rate, _sanitize_terminal_growth_rate, _scenario_assumptions,
    _forecast_map, _scenario_first_projection_forecast, _infer_revenue_growth_rate
)

def _set_comment(worksheet: Worksheet, cell_ref: str, text: str, *, author: str = "DCF Builder") -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    if not text.strip():
        return
    cell.comment = Comment(text, author)

def _safe_set(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    _safe_set_with_options(worksheet, cell_ref, value, clear_if_none=False)

def _safe_set_or_clear(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    _safe_set_with_options(worksheet, cell_ref, value, clear_if_none=True)

def _safe_set_with_options(
    worksheet: Worksheet,
    cell_ref: str,
    value: Any,
    *,
    clear_if_none: bool,
) -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    if cell.data_type == "f":
        return
    if value is None:
        if not clear_if_none:
            return
        cell.value = None
        cell.hyperlink = None
        return
    cell.value = value

def _force_set(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.value = value

def _axis_from_bounds(
    values: list[float],
    *,
    center: float,
    step: float,
    min_value: float,
    max_value: float,
) -> list[float]:
    if len(values) >= 2:
        low = max(min_value, min(values))
        high = min(max_value, max(values))
        if high <= low:
            high = min(max_value, low + step * 4)
        if high > low:
            return [round(low + (high - low) * idx / 4, 4) for idx in range(5)]

    seed = center if center == center else (min_value + max_value) / 2
    axis = [seed + (idx - 2) * step for idx in range(5)]
    clamped = [max(min_value, min(max_value, value)) for value in axis]
    for idx in range(1, len(clamped)):
        if clamped[idx] <= clamped[idx - 1]:
            clamped[idx] = min(max_value, clamped[idx - 1] + max(step / 2, 0.0005))
    return [round(value, 4) for value in clamped]

def _uniform_axis_step(values: list[float], *, tolerance: float = 1e-6) -> float | None:
    if len(values) != 5:
        return None
    deltas = [values[idx + 1] - values[idx] for idx in range(len(values) - 1)]
    if any(delta <= 0 for delta in deltas):
        return None
    first = deltas[0]
    if all(abs(delta - first) <= tolerance for delta in deltas[1:]):
        return first
    return None

def _set_percent_axis_row(
    sheet: Worksheet,
    *,
    cells: tuple[str, str, str, str, str],
    values: list[float],
) -> None:
    step = _uniform_axis_step(values)
    if step is not None:
        center = cells[2]
        _force_set(sheet, center, values[2])
        _force_set(sheet, cells[1], f"={center}-{step:.4f}")
        _force_set(sheet, cells[0], f"={cells[1]}-{step:.4f}")
        _force_set(sheet, cells[3], f"={center}+{step:.4f}")
        _force_set(sheet, cells[4], f"={cells[3]}+{step:.4f}")
    else:
        for cell_ref, value in zip(cells, values, strict=False):
            _force_set(sheet, cell_ref, value)

    for cell_ref in cells:
        sheet[cell_ref].number_format = "0.0%"

def _set_percent_axis_column(
    sheet: Worksheet,
    *,
    cells: tuple[str, str, str, str, str],
    values: list[float],
) -> None:
    step = _uniform_axis_step(values)
    if step is not None:
        center = cells[2]
        _force_set(sheet, center, values[2])
        _force_set(sheet, cells[1], f"={center}-{step:.4f}")
        _force_set(sheet, cells[0], f"={cells[1]}-{step:.4f}")
        _force_set(sheet, cells[3], f"={center}+{step:.4f}")
        _force_set(sheet, cells[4], f"={cells[3]}+{step:.4f}")
    else:
        for cell_ref, value in zip(cells, values, strict=False):
            _force_set(sheet, cell_ref, value)

    for cell_ref in cells:
        sheet[cell_ref].number_format = "0.0%"

def _clear_sensitivity_blocks(*scenario_sheets: Worksheet) -> None:
    for sheet in scenario_sheets:
        for row in range(117, 131):
            for col in "BCDEFGHIJKLMNO":
                _force_set(sheet, f"{col}{row}", None)

def _map_scenario_forecasts_to_sheet(
    sheet: Worksheet,
    forecast_by_year: dict[int, dict[str, Any]],
    timeline_years: list[int],
    divisor: float,
) -> None:
    projection_columns = DCF_TIMELINE_COLUMNS[2:]  # J..Q
    projection_years = timeline_years[2:] if len(timeline_years) >= 3 else []
    first_projection_year = next((year for year in projection_years if year in forecast_by_year), None)

    for idx, col in enumerate(projection_columns):
        if idx >= len(projection_years):
            break
        projection_year = projection_years[idx]
        forecast = forecast_by_year.get(projection_year)
        if not isinstance(forecast, dict):
            continue

        revenue = _to_float(forecast.get("revenue"))
        if revenue is not None and projection_year == first_projection_year:
            _force_set(sheet, f"{col}20", _scale(revenue, divisor))

def _apply_scenario_snapshot_to_sheet(
    sheet: Worksheet,
    snapshot: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    assumptions = _scenario_assumptions(snapshot)
    forecast_by_year = _forecast_map(snapshot)

    tax_rate = _to_float(assumptions.get("taxRate"))
    wacc_rate = _sanitize_wacc_rate(assumptions.get("waccRate"))
    da_pct_revenue = _to_float(assumptions.get("daPctRevenue"))
    revenue_growth_rate = _to_float(assumptions.get("revenueGrowthRate"))
    if revenue_growth_rate is None:
        revenue_growth_rate = _to_float(assumptions.get("revenueGrowth"))
    if revenue_growth_rate is None:
        revenue_growth_rate = _infer_revenue_growth_rate(snapshot.get("forecasts") if isinstance(snapshot.get("forecasts"), list) else [])
    terminal_growth = _sanitize_terminal_growth_rate(
        assumptions.get("terminalGrowthRate"),
        reference_wacc=wacc_rate,
    )
    exit_multiple = _to_float(assumptions.get("terminalExitMultiple"))

    if tax_rate is not None:
        _force_set(sheet, "F11", tax_rate)
    if da_pct_revenue is not None:
        _force_set(sheet, "F13", da_pct_revenue)
    if revenue_growth_rate is not None:
        _force_set(sheet, "F14", revenue_growth_rate)
        sheet["F14"].number_format = "0.0%"
    if terminal_growth is not None:
        _force_set(sheet, "Q103", terminal_growth)
    if exit_multiple is not None:
        _force_set(sheet, "C16", exit_multiple)

    first_projection = _scenario_first_projection_forecast(forecast_by_year, timeline_years)
    if isinstance(first_projection, dict):
        capex = _to_float(first_projection.get("capex"))
        nwc_change = _first_float(first_projection, "nwcChange", "nwc_change")
        if capex is not None:
            _force_set(sheet, "F9", _scale(abs(capex), divisor))
        if nwc_change is not None:
            _force_set(sheet, "F10", _scale(abs(nwc_change), divisor))

    _map_scenario_forecasts_to_sheet(sheet, forecast_by_year, timeline_years, divisor)

def _link_dcf_income_statement_to_recalculated_data(*scenario_sheets: Worksheet) -> None:
    for sheet in scenario_sheets:
        for idx, dcf_col in enumerate(DCF_TIMELINE_COLUMNS):
            recalc_col = RECALC_COLUMNS[idx]
            _force_set(sheet, f"{dcf_col}20", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}12")
            _force_set(sheet, f"{dcf_col}24", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}16")
            _force_set(sheet, f"{dcf_col}27", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}17")
            _force_set(sheet, f"{dcf_col}36", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}24")
            _force_set(sheet, f"{dcf_col}39", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}25")
            _force_set(sheet, f"{dcf_col}42", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}26")
            _force_set(sheet, f"{dcf_col}45", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}27")

def _normalize_public_dcf_assumption_block(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    market = payload.get("market", {})
    cash = _to_float(market.get("cash")) or 0.0
    debt = _to_float(market.get("debt")) or 0.0
    non_operating_assets = _to_float(market.get("nonOperatingAssets")) or 0.0

    for sheet, wacc_cell in (
        (dcf_base, "D34"),
        (dcf_bull, "D35"),
        (dcf_bear, "D36"),
    ):
        _safe_set(sheet, "B8", "Valuation Inputs")
        _safe_set(sheet, "B9", "Enterprise Value")
        _safe_set(sheet, "B10", "EV / EBITDA (LTM)")
        _safe_set(sheet, "B11", "Current Equity Value")
        _safe_set(sheet, "B12", "Implied Equity Value")
        _safe_set(sheet, "E14", "Revenue Growth Rate")
        _safe_set(sheet, "B15", "Terminal Assumptions")
        _safe_set(sheet, "B16", "Exit EBITDA Multiple")
        _safe_set(sheet, "B17", "Cash and Cash Equivalents")
        _safe_set(sheet, "B18", "Income Statement")
        _safe_set_or_clear(sheet, "B19", None)
        _safe_set_or_clear(sheet, "B13", None)
        _safe_set_or_clear(sheet, "C13", None)
        _safe_set_or_clear(sheet, "B14", None)
        _safe_set_or_clear(sheet, "C14", None)
        sheet["F14"]._style = copy(sheet["F11"]._style)
        _safe_set_or_clear(sheet, "F14", _to_float(sheet["F14"].value) or _to_float(sheet["C14"].value) or _to_float(sheet["C13"].value))
        sheet["F14"].number_format = "0.0%"
        _safe_set_or_clear(sheet, "E15", None)
        _safe_set_or_clear(sheet, "E16", None)
        _safe_set_or_clear(sheet, "E19", None)
        _force_set(sheet, "C17", _scale(cash, divisor))
        _force_set(sheet, "C10", "=IFERROR(C9/L54,0)")
        _force_set(sheet, "C12", f"=C9-{DCF_HELPER_DEBT_CELL}+{DCF_HELPER_CASH_CELL}+{DCF_HELPER_NON_OP_CELL}")
        _force_set(sheet, "F12", f"=WACC!{wacc_cell}")
        _safe_set_or_clear(sheet, "F16", None)
        _safe_set_or_clear(sheet, "F17", None)
        _safe_set_or_clear(sheet, "F18", None)
        _safe_set_or_clear(sheet, "F19", None)
        _force_set(sheet, DCF_HELPER_CASH_CELL, "=C17")
        _force_set(sheet, DCF_HELPER_DEBT_CELL, _scale(debt, divisor))
        _force_set(sheet, DCF_HELPER_NON_OP_CELL, _scale(non_operating_assets, divisor))
        for assumption_ref in ("F9", "F10", "F11", "F12", "F13", "F14", "C16"):
            sheet[assumption_ref].comment = None

def _enforce_core_public_dcf_formulas(*scenario_sheets: Worksheet) -> None:
    projection_columns = DCF_TIMELINE_COLUMNS[2:]  # J..Q
    formula_projection_columns = DCF_TIMELINE_COLUMNS[5:]  # M..Q

    for sheet in scenario_sheets:
        for col in DCF_TIMELINE_COLUMNS:
            _force_set(sheet, f"{col}30", f"={col}24+{col}27")
            _force_set(sheet, f"{col}32", f"={col}20-{col}30")
            _force_set(sheet, f"{col}33", f"=IFERROR({col}32/{col}20,0)")
            _force_set(sheet, f"{col}54", f"={col}51+{col}68")
            _force_set(sheet, f"{col}55", f"=IFERROR({col}54/{col}20,0)")
            _force_set(sheet, f"{col}57", f"=-{col}51*{col}58")
            _force_set(sheet, f"{col}58", "=$F$11")
            _force_set(sheet, f"{col}60", f"={col}51+{col}57")
            _force_set(sheet, f"{col}68", f"={col}69*{col}20")
            _force_set(sheet, f"{col}70", f"=IFERROR(-{col}68/{col}65,0)")
            _force_set(sheet, f"{col}74", f"={col}60")
            _force_set(sheet, f"{col}75", f"={col}68")
            _force_set(sheet, f"{col}76", f"={col}65")
            _force_set(sheet, f"{col}78", f"=SUM({col}74:{col}77)")
            _force_set(sheet, f"{col}79", f"=IFERROR({col}78/{col}20,0)")
            _force_set(sheet, f"{col}85", "=$F$12")
            if col not in formula_projection_columns:
                _force_set(sheet, f"{col}25", f"=IFERROR({col}24/{col}20,0)")
                _force_set(sheet, f"{col}37", f"=IFERROR({col}36/{col}20,0)")
                _force_set(sheet, f"{col}40", f"=IFERROR({col}39/{col}20,0)")
                _force_set(sheet, f"{col}43", f"=IFERROR({col}42/{col}20,0)")
                _force_set(sheet, f"{col}46", f"=IFERROR({col}45/{col}20,0)")

        for idx, col in enumerate(formula_projection_columns):
            prev_col = DCF_TIMELINE_COLUMNS[5 + idx - 1]
            _force_set(sheet, f"{col}25", f"={prev_col}25")
            _force_set(sheet, f"{col}37", f"={prev_col}37")
            _force_set(sheet, f"{col}40", f"={prev_col}40")
            _force_set(sheet, f"{col}43", f"={prev_col}43")
            _force_set(sheet, f"{col}46", f"={prev_col}46")
            _force_set(sheet, f"{col}24", f"={col}20*{col}25")
            _force_set(sheet, f"{col}36", f"={col}20*{col}37")
            _force_set(sheet, f"{col}39", f"={col}20*{col}40")
            _force_set(sheet, f"{col}42", f"={col}20*$F$13")
            _force_set(sheet, f"{col}45", f"={col}20*{col}46")
            _force_set(sheet, f"{col}65", "=-$F$9")
            _force_set(sheet, f"{col}77", "=-$F$10")

        _force_set(sheet, "E81", "=I9")
        for idx, col in enumerate(projection_columns):
            prev_col = projection_columns[idx - 1] if idx > 0 else None
            _force_set(sheet, f"{col}82", "=I11" if col == "J" else f"=EOMONTH({prev_col}82,12)")
            _force_set(sheet, f"{col}83", f"=({col}82-E81)/365" if col == "J" else f"={prev_col}83+1")
            _force_set(sheet, f"{col}84", f"={col}83/2" if col == "J" else f"={col}83-0.5")
            _force_set(sheet, f"{col}87", f"={col}78/(1+{col}85)^{col}84")

        _force_set(sheet, "Q92", "=Q54*$C$16")
        _force_set(sheet, "Q93", "=Q92/(1+Q85)^Q83")
        _force_set(sheet, "Q94", "=Q93+SUM(J87:Q87)")
        _force_set(sheet, "C9", "=Q94")
        _force_set(sheet, "Q95", f"=-{DCF_HELPER_DEBT_CELL_ABS}")
        _force_set(sheet, "Q96", f"={DCF_HELPER_NON_OP_CELL_ABS}")
        _force_set(sheet, "Q97", f"={DCF_HELPER_CASH_CELL_ABS}")
        _force_set(sheet, "Q98", "=SUM(Q94:Q97)")
        _force_set(sheet, "Q99", "=C11")
        _force_set(sheet, "Q100", "=IFERROR(Q98/Q99 - 1,0)")
        _force_set(sheet, "Q104", "=Q78*(1+Q103)")
        _force_set(sheet, "Q105", "=IFERROR(IF(Q85>Q103,Q104/(Q85-Q103),0),0)")
        _force_set(sheet, "Q106", "=Q105/(1+Q85)^Q83")
        _force_set(sheet, "Q107", "=Q106+SUM(J87:Q87)")
        _force_set(sheet, "Q108", f"=-{DCF_HELPER_DEBT_CELL_ABS}")
        _force_set(sheet, "Q109", f"={DCF_HELPER_NON_OP_CELL_ABS}")
        _force_set(sheet, "Q110", f"={DCF_HELPER_CASH_CELL_ABS}")
        _force_set(sheet, "Q111", "=SUM(Q107:Q110)")
        _force_set(sheet, "Q112", "=C11")
        _force_set(sheet, "Q113", "=IFERROR(Q111/Q112 - 1,0)")

        first_forecast_idx = next(
            (idx for idx, col in enumerate(DCF_TIMELINE_COLUMNS) if isinstance(sheet[f"{col}18"].value, str) and str(sheet[f"{col}18"].value).endswith("E")),
            None,
        )
        if first_forecast_idx is not None:
            anchor_col = DCF_TIMELINE_COLUMNS[first_forecast_idx]
            sheet[f"{anchor_col}20"].comment = None
            for idx in range(first_forecast_idx + 1, len(DCF_TIMELINE_COLUMNS)):
                col = DCF_TIMELINE_COLUMNS[idx]
                prev_col = DCF_TIMELINE_COLUMNS[idx - 1]
                _force_set(sheet, f"{col}20", f"=IFERROR({prev_col}20*(1+$F$14),{prev_col}20)")

def _scenario_choose_formula(base_cell: str, bull_cell: str | None = None, bear_cell: str | None = None) -> str:
    bull_ref = bull_cell or base_cell
    bear_ref = bear_cell or base_cell
    return (
        "=CHOOSE(Cover!$C$12,"
        f"'{SHEET_DCF_BASE}'!{base_cell},"
        f"'{SHEET_DCF_BULL}'!{bull_ref},"
        f"'{SHEET_DCF_BEAR}'!{bear_ref})"
    )

def _enforce_outputs_bridge_formulas(outputs: Worksheet) -> None:
    _force_set(outputs, "E18", _scenario_choose_formula("I9"))
    _force_set(outputs, "J19", _scenario_choose_formula("I11"))
    _force_set(outputs, "J20", "=(J19-E18)/365")
    _force_set(outputs, "J21", "=J20/2")
    _force_set(outputs, "D27", _scenario_choose_formula("$F$12"))
    _force_set(outputs, "H27", "=D27")
    _force_set(outputs, "D28", _scenario_choose_formula("$C$16"))
    _force_set(outputs, "H28", _scenario_choose_formula("$Q$103"))

    for col in ("H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"):
        _force_set(outputs, f"{col}8", _scenario_choose_formula(f"{col}51"))
        _force_set(outputs, f"{col}9", _scenario_choose_formula(f"{col}68"))
        _force_set(outputs, f"{col}10", _scenario_choose_formula(f"{col}54"))
        _force_set(outputs, f"{col}12", _scenario_choose_formula(f"{col}60"))
        _force_set(outputs, f"{col}13", f"={col}9")
        _force_set(outputs, f"{col}14", _scenario_choose_formula(f"{col}60"))
        _force_set(outputs, f"{col}15", f"=-{_scenario_choose_formula('$F$10')[1:]}")
        _force_set(outputs, f"{col}22", _scenario_choose_formula("$F$12"))

    _safe_set(outputs, "B37", "(+) Non-Operating Assets")
    _safe_set(outputs, "F37", "(+) Non-Operating Assets")
    _force_set(outputs, "D33", "=D32/(1+D27)^Q20")
    _force_set(outputs, "H33", "=H32/(1+H27)^Q20")
    _force_set(outputs, "D36", f"=-{_scenario_choose_formula(DCF_HELPER_DEBT_CELL_ABS)[1:]}")
    _force_set(outputs, "H36", "=D36")
    _force_set(outputs, "D37", _scenario_choose_formula(DCF_HELPER_NON_OP_CELL_ABS))
    _force_set(outputs, "H37", "=D37")
    _force_set(outputs, "D38", _scenario_choose_formula(DCF_HELPER_CASH_CELL_ABS))
    _force_set(outputs, "H38", "=D38")
    _force_set(outputs, "D41", _scenario_choose_formula("$C$11"))
    _force_set(outputs, "H41", "=D41")

def _rewrite_formula_sheet_name_references(workbook: Workbook, *, old_name: str, new_name: str) -> None:
    old_ref = f"'{old_name}'!"
    new_ref = f"'{new_name}'!"
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and old_ref in value:
                    cell.value = value.replace(old_ref, new_ref)
                hyperlink = cell.hyperlink
                if hyperlink is not None and isinstance(hyperlink.location, str) and old_ref in hyperlink.location:
                    hyperlink.location = hyperlink.location.replace(old_ref, new_ref)
