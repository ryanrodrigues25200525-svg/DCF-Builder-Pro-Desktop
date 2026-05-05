from __future__ import annotations
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from openpyxl import load_workbook
from defusedxml import ElementTree as ET
from copy import copy

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

def _sheet_name_by_path(archive: ZipFile) -> dict[str, str]:
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    ns_workbook = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rels = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    rel_target_by_id: dict[str, str] = {}
    for rel in rels_xml.findall("r:Relationship", ns_rels):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if not rel_id or not target:
            continue
        rel_target_by_id[rel_id] = _normalize_rel_target(target)

    sheet_name_by_path: dict[str, str] = {}
    for sheet in workbook_xml.findall("x:sheets/x:sheet", ns_workbook):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not name or not rel_id:
            continue
        target = rel_target_by_id.get(rel_id)
        if target:
            sheet_name_by_path[target] = name

    return sheet_name_by_path

def _normalize_rel_target(target: str) -> str:
    normalized = target.strip().replace("\\", "/")
    if normalized.startswith("/"):
        normalized = normalized[1:]
    while normalized.startswith("./"):
        normalized = normalized[2:]
    if not normalized.startswith("xl/"):
        normalized = f"xl/{normalized}"
    return normalized

def _template_cell_styles_by_sheet(
    template_zip: ZipFile,
    template_sheet_path_by_name: dict[str, str],
) -> dict[str, dict[str, str]]:
    ns = {"x": _NS_MAIN}
    out: dict[str, dict[str, str]] = {}
    for sheet_name, sheet_path in template_sheet_path_by_name.items():
        if sheet_path not in template_zip.namelist():
            continue
        root = ET.fromstring(template_zip.read(sheet_path))
        styles: dict[str, str] = {}
        for cell in root.findall(".//x:c", ns):
            address = cell.attrib.get("r")
            style_id = cell.attrib.get("s")
            if not address or style_id is None:
                continue
            styles[address] = style_id
        out[sheet_name] = styles
    return out

def _finalize_output_year_labels(workbook_bytes: bytes) -> bytes:
    wb = load_workbook(filename=BytesIO(workbook_bytes), data_only=False)
    if "Outputs - Base" not in wb.sheetnames or "DCF Model - Base (1)" not in wb.sheetnames:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    outputs = wb["Outputs - Base"]
    base = wb["DCF Model - Base (1)"]
    dcf_sheets = [wb[name] for name in ("DCF Model - Base (1)", "DCF Model - Bull (2)", "DCF Model - Bear (3)") if name in wb.sheetnames]
    for col in "HIJKLMNOPQ":
        label = base[f"{col}18"].value
        if not (isinstance(label, str) and label.startswith("FY")):
            continue
        outputs[f"{col}6"].value = label
        outputs[f"{col}6"].number_format = "@"
        for sheet in dcf_sheets:
            sheet[f"{col}89"].value = label
            sheet[f"{col}89"].number_format = "@"

    g_mirror_rows = (
        18, 20, 21, 24, 25, 27, 28, 30, 32, 33, 36, 37, 39, 40, 42, 43, 45, 46,
        48, 49, 51, 52, 54, 55, 57, 58, 60, 61, 63, 65, 66, 67, 68, 69, 70, 72,
        74, 75, 76, 77, 78, 79,
    )
    for sheet in dcf_sheets:
        if sheet["G18"].value in (None, ""):
            continue
        for row in g_mirror_rows:
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
